"""Export du catalogue en classeur Excel (.xlsx) multi-feuilles.

Chaque table clé de la base (fiches, catégories, plateformes, tarifs,
sources, historique) devient une feuille avec en-têtes figés, filtres
automatiques et mise en forme façon tableau Excel. Une feuille "Tableau
de bord" résume les volumes avec un graphique.
"""

from __future__ import annotations

import sqlite3
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from ia_database.icons import favicon_url

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9", showRowStripes=True, showFirstColumn=False, showLastColumn=False
)
ICON_SIZE = 20


MAX_ICON_DOWNLOADS = 500  # nombre de domaines uniques, pas de lignes (cf. cache ci-dessous)
MAX_EMBEDDED_ROWS = 300  # au-delà, le fichier devient lourd (images non dédupliquées par Excel)


def _embed_icons(worksheet: Worksheet, official_urls: list[str | None], insecure: bool) -> tuple[int, int]:
    """Insère une miniature de favicon en colonne A pour chaque ligne, au mieux.

    Un même domaine (ex. huggingface.co pour des milliers de modèles) partage
    une seule requête réseau grâce à un cache par domaine — indispensable dès
    que le catalogue dépasse quelques centaines de fiches, sinon une collecte
    de plusieurs milliers de modèles prendrait des heures.

    Best-effort : réseau indisponible, favicon absent ou format illisible ne
    doivent jamais faire échouer l'export, juste laisser la cellule vide.
    Retourne (tentatives, réussites) pour permettre d'avertir l'appelant si
    tout a échoué (typiquement : TLS intercepté par un antivirus, cf. --insecure).
    """
    try:
        import requests
        from PIL import Image as PILImage
    except ImportError:
        return (0, 0)

    if insecure:
        import urllib3

        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    cache: dict[str, bytes | None] = {}
    attempted = 0
    succeeded = 0
    consecutive_failures = 0
    for offset, url in enumerate(official_urls, start=2):
        if consecutive_failures >= 5:
            break
        source_url = favicon_url(url, size=64)
        if not source_url:
            continue

        if source_url not in cache:
            if len(cache) >= MAX_ICON_DOWNLOADS:
                continue
            attempted += 1
            try:
                response = requests.get(source_url, timeout=4, verify=not insecure)
                response.raise_for_status()
                image = PILImage.open(BytesIO(response.content)).convert("RGBA")
                image.thumbnail((ICON_SIZE, ICON_SIZE))
                buffer = BytesIO()
                image.save(buffer, format="PNG")
                cache[source_url] = buffer.getvalue()
                consecutive_failures = 0
            except Exception:  # noqa: BLE001 - icône manquante, jamais bloquant
                cache[source_url] = None
                consecutive_failures += 1
                continue

        png_bytes = cache[source_url]
        if png_bytes is None:
            continue
        xl_image = XLImage(BytesIO(png_bytes))
        with PILImage.open(BytesIO(png_bytes)) as sized:
            xl_image.width, xl_image.height = sized.size
        worksheet.add_image(xl_image, f"A{offset}")
        worksheet.row_dimensions[offset].height = max(worksheet.row_dimensions[offset].height or 15, 16)
        succeeded += 1
        if succeeded >= MAX_EMBEDDED_ROWS:
            break
    return (attempted, succeeded)


def _write_table(worksheet: Worksheet, name: str, columns: list[str], rows: list[tuple]) -> None:
    worksheet.append(columns)
    for row in rows:
        worksheet.append(row)

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    worksheet.freeze_panes = "A2"

    last_row = max(worksheet.max_row, 1)
    last_column = get_column_letter(len(columns))
    if last_row > 1:
        table = Table(displayName=name, ref=f"A1:{last_column}{last_row}")
        table.tableStyleInfo = TABLE_STYLE
        worksheet.add_table(table)

    for index, column in enumerate(columns, start=1):
        width = max(len(str(column)), *(len(str(row[index - 1])) for row in rows)) if rows else len(str(column))
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 60)


def _fetch(connection: sqlite3.Connection, sql: str) -> tuple[list[str], list[tuple]]:
    cursor = connection.execute(sql)
    columns = [description[0] for description in cursor.description]
    rows = [tuple(row) for row in cursor.fetchall()]
    return columns, rows


def export_workbook(
    database_path: Path,
    output_path: Path,
    with_icons: bool = True,
    insecure: bool = False,
) -> None:
    """Génère le classeur Excel à partir de la base SQLite locale.

    `with_icons` télécharge une miniature de favicon par fiche (best-effort,
    désactivable avec `--no-icons` si le poste est hors ligne). `insecure`
    désactive la vérification TLS de ces téléchargements (réseau dont
    l'inspection HTTPS casse la vérification, ex. antivirus).
    """
    workbook = Workbook()
    workbook.remove(workbook.active)

    with sqlite3.connect(database_path) as connection:
        catalogue_columns, catalogue_rows = _fetch(
            connection,
            """
            SELECT
                item.slug, item.name, item.item_type, company.name AS editeur,
                item.official_url, item.api_available, item.works_offline, item.is_open_source,
                item.status, item.downloads, item.likes,
                COALESCE(GROUP_CONCAT(DISTINCT category.name), '') AS categories,
                COALESCE(GROUP_CONCAT(DISTINCT platform.name), '') AS plateformes,
                COALESCE(GROUP_CONCAT(DISTINCT tag.name), '') AS tags
            FROM catalog_items AS item
            LEFT JOIN companies AS company ON company.id = item.company_id
            LEFT JOIN item_categories AS ic ON ic.item_id = item.id
            LEFT JOIN categories AS category ON category.id = ic.category_id
            LEFT JOIN item_platforms AS ip ON ip.item_id = item.id
            LEFT JOIN platforms AS platform ON platform.id = ip.platform_id
            LEFT JOIN item_tags AS it ON it.item_id = item.id
            LEFT JOIN tags AS tag ON tag.id = it.tag_id
            GROUP BY item.id
            ORDER BY (COALESCE(item.downloads, 0) + COALESCE(item.likes, 0) * 20) DESC, item.name COLLATE NOCASE
            """,
        )
        official_urls = [row[catalogue_columns.index("official_url")] for row in catalogue_rows]
        catalogue_columns = ["icone", *catalogue_columns]
        catalogue_rows = [("", *row) for row in catalogue_rows]

        category_columns, category_rows = _fetch(
            connection,
            """
            SELECT category.slug, category.name, category.description,
                   COUNT(ic.item_id) AS nb_fiches
            FROM categories AS category
            LEFT JOIN item_categories AS ic ON ic.category_id = category.id
            GROUP BY category.id
            ORDER BY nb_fiches DESC, category.name
            """,
        )

        platform_columns, platform_rows = _fetch(
            connection,
            """
            SELECT platform.slug, platform.name, COUNT(ip.item_id) AS nb_fiches
            FROM platforms AS platform
            LEFT JOIN item_platforms AS ip ON ip.platform_id = platform.id
            GROUP BY platform.id
            ORDER BY nb_fiches DESC, platform.name
            """,
        )

        pricing_columns, pricing_rows = _fetch(
            connection,
            """
            SELECT item.name AS fiche, pricing.plan_name, pricing.pricing_type,
                   pricing.price_amount, pricing.currency, pricing.billing_period, pricing.url
            FROM pricing_plans AS pricing
            JOIN catalog_items AS item ON item.id = pricing.item_id
            ORDER BY item.name COLLATE NOCASE, pricing.plan_name
            """,
        )

        source_columns, source_rows = _fetch(
            connection,
            """
            SELECT source.slug, source.name, source.base_url, source.is_active,
                   COUNT(record.id) AS nb_fiches_liees, MAX(record.retrieved_at) AS derniere_collecte
            FROM sources AS source
            LEFT JOIN source_records AS record ON record.source_id = source.id
            GROUP BY source.id
            ORDER BY nb_fiches_liees DESC, source.name
            """,
        )

        history_columns, history_rows = _fetch(
            connection,
            """
            SELECT item.name AS fiche, history.field_name, history.old_value,
                   history.new_value, history.changed_at
            FROM item_history AS history
            JOIN catalog_items AS item ON item.id = history.item_id
            ORDER BY history.changed_at DESC
            """,
        )

        tag_columns, tag_rows = _fetch(
            connection,
            """
            SELECT tag.slug, tag.name, COUNT(it.item_id) AS nb_fiches
            FROM tags AS tag
            LEFT JOIN item_tags AS it ON it.tag_id = tag.id
            GROUP BY tag.id
            ORDER BY nb_fiches DESC, tag.name
            """,
        )

        license_columns, license_rows = _fetch(
            connection,
            """
            SELECT license.name, license.spdx_id, license.url,
                   COALESCE(GROUP_CONCAT(DISTINCT item.name), '') AS fiches
            FROM licenses AS license
            LEFT JOIN item_licenses AS il ON il.license_id = license.id
            LEFT JOIN catalog_items AS item ON item.id = il.item_id
            GROUP BY license.id
            ORDER BY license.name
            """,
        )

        total_items = connection.execute("SELECT COUNT(*) FROM catalog_items").fetchone()[0]
        open_source_count = connection.execute(
            "SELECT COUNT(*) FROM catalog_items WHERE is_open_source = 1"
        ).fetchone()[0]
        offline_count = connection.execute(
            "SELECT COUNT(*) FROM catalog_items WHERE works_offline = 1"
        ).fetchone()[0]
        api_count = connection.execute(
            "SELECT COUNT(*) FROM catalog_items WHERE api_available = 1"
        ).fetchone()[0]

    sheets = [
        ("Catalogue", catalogue_columns, catalogue_rows),
        ("Categories", category_columns, category_rows),
        ("Plateformes", platform_columns, platform_rows),
        ("Tags", tag_columns, tag_rows),
        ("Tarifs", pricing_columns, pricing_rows),
        ("Licences", license_columns, license_rows),
        ("Sources", source_columns, source_rows),
        ("Historique", history_columns, history_rows),
    ]
    for sheet_name, columns, rows in sheets:
        worksheet = workbook.create_sheet(sheet_name)
        _write_table(worksheet, f"tbl_{sheet_name}", columns, rows)
        if sheet_name == "Catalogue":
            worksheet.column_dimensions["A"].width = 4
            if with_icons:
                attempted, succeeded = _embed_icons(worksheet, official_urls, insecure)
                if attempted and not succeeded:
                    print(
                        "Avertissement : aucune icône n'a pu être téléchargée "
                        f"({attempted} tentatives). Si le réseau est filtré par un antivirus "
                        "ou un proxy qui intercepte le TLS, relance avec --insecure. "
                        "Sinon, utilise --no-icons pour un export sans avertissement."
                    )
                elif succeeded >= MAX_EMBEDDED_ROWS < len(official_urls):
                    print(
                        f"Icônes limitées aux {MAX_EMBEDDED_ROWS} premières fiches "
                        f"(sur {len(official_urls)}) pour garder le classeur léger."
                    )

    dashboard = workbook.create_sheet("Tableau de bord", 0)
    dashboard["A1"] = "Catalogue IA — indicateurs"
    dashboard["A1"].font = Font(bold=True, size=14)
    kpis = [
        ("Fiches au total", total_items),
        ("Disponibles via API", api_count),
        ("Utilisables hors ligne", offline_count),
        ("Open source", open_source_count),
    ]
    for offset, (label, value) in enumerate(kpis, start=3):
        dashboard[f"A{offset}"] = label
        dashboard[f"B{offset}"] = value
        dashboard[f"A{offset}"].font = Font(bold=True)

    chart_start_row = 9
    dashboard[f"A{chart_start_row - 1}"] = "Fiches par catégorie"
    dashboard[f"A{chart_start_row - 1}"].font = Font(bold=True)
    dashboard.append([])
    for row_index, (slug, name, _description, count) in enumerate(category_rows, start=chart_start_row):
        dashboard[f"A{row_index}"] = name
        dashboard[f"B{row_index}"] = count

    if category_rows:
        chart = BarChart()
        chart.title = "Fiches par catégorie"
        chart.y_axis.title = "Nombre de fiches"
        last_row = chart_start_row + len(category_rows) - 1
        data = Reference(dashboard, min_col=2, min_row=chart_start_row, max_row=last_row)
        categories = Reference(dashboard, min_col=1, min_row=chart_start_row, max_row=last_row)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(categories)
        chart.legend = None
        dashboard.add_chart(chart, f"D{chart_start_row}")

    dashboard.column_dimensions["A"].width = 32
    dashboard.column_dimensions["B"].width = 16

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
