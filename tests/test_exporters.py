from __future__ import annotations

import sqlite3
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import load_workbook

from ia_database.exporters.xlsx import export_workbook

ROOT = Path(__file__).resolve().parents[1]


def _seeded_database(tmp_path: Path) -> Path:
    database_path = tmp_path / "catalog.db"
    schema = (ROOT / "database" / "schema.sql").read_text(encoding="utf-8")
    seed = (ROOT / "database" / "seed.sql").read_text(encoding="utf-8")
    with sqlite3.connect(database_path) as connection:
        connection.executescript(schema)
        connection.executescript(seed)
    return database_path


def test_export_workbook_has_expected_sheets(tmp_path: Path) -> None:
    database_path = _seeded_database(tmp_path)

    output_path = tmp_path / "catalogue_ia.xlsx"
    export_workbook(database_path, output_path, with_icons=False)

    assert output_path.exists()
    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [
        "Tableau de bord",
        "Catalogue",
        "Categories",
        "Plateformes",
        "Tags",
        "Tarifs",
        "Licences",
        "Sources",
        "Historique",
    ]

    catalogue = workbook["Catalogue"]
    assert catalogue["A1"].value == "icone"
    assert catalogue["B1"].value == "slug"
    assert catalogue.max_row == 52  # en-tête + 51 fiches de démarrage

    dashboard = workbook["Tableau de bord"]
    assert dashboard["B3"].value == 51


def test_export_workbook_embeds_icons_best_effort(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    database_path = _seeded_database(tmp_path)

    # Un petit PNG 1x1 valide, généré à la volée pour éviter tout vrai appel réseau.
    from PIL import Image as PILImage

    buffer = BytesIO()
    PILImage.new("RGBA", (1, 1), (255, 0, 0, 255)).save(buffer, format="PNG")
    tiny_png = buffer.getvalue()

    class FakeResponse:
        content = tiny_png

        def raise_for_status(self) -> None:
            return None

    import requests

    monkeypatch.setattr(requests, "get", lambda *a, **k: FakeResponse())

    output_path = tmp_path / "catalogue_with_icons.xlsx"
    export_workbook(database_path, output_path, with_icons=True)

    workbook = load_workbook(output_path)
    catalogue = workbook["Catalogue"]
    assert len(catalogue._images) == 51
